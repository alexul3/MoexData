import openpyxl
from transformers import pipeline

model = pipeline("sentiment-analysis", model="cointegrated/rubert-tiny-sentiment-balanced")

wb = openpyxl.load_workbook('news.xlsx')
sheet = wb.active

for row in range(1, sheet.max_row + 1):
    text = sheet.cell(row=row, column=1).value
    if text and isinstance(text, str):
        if len(text) > 512:
            text = text[:512]
        result = model(text)[0]
        label = result['label']
        score = result['score']
        if label == 'positive':
            sentiment = score
        elif label == 'negative':
            sentiment = -score
        else:
            sentiment = 0.0
        sheet.cell(row=row, column=2, value=sentiment)
    else:
        sheet.cell(row=row, column=2, value=0.0)

wb.save('news_with_sentiment.xlsx')